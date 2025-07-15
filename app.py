from flask import Flask, request, jsonify, render_template_string, send_file
import json
from datetime import datetime
import os
import tempfile
import io
import base64
from PIL import Image
import pytesseract
from docx import Document
import re

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

class SimpleTimesheetProcessor:
    def __init__(self):
        # Set tesseract path for Windows
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        self.timesheet_data = []

    def extract_images_from_word_file(self, word_file_path):
        """Extract images from Word document"""
        images = []
        try:
            doc = Document(word_file_path)
            for rel in doc.part.rels.values():
                if "image" in rel.target_ref:
                    try:
                        image_data = rel.target_part.blob
                        image = Image.open(io.BytesIO(image_data))
                        if image.mode != 'RGB':
                            image = image.convert('RGB')
                        images.append(image)
                    except:
                        continue
        except Exception as e:
            print(f"Error with document: {e}")
        return images

    def extract_name_from_filename(self, filename):
        """Extract name from filename"""
        name = filename.replace('.docx', '').replace('.doc', '')
        name = re.sub(r'-\d{4}', '', name)  # Remove year
        name = re.sub(r'[_-]', ' ', name)   # Replace underscores/dashes with spaces
        name = name.strip()
        return name if name else "Unknown"

    def extract_text_from_image(self, image):
        """Extract text using OCR"""
        try:
            config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
            text = pytesseract.image_to_string(image, config=config)
            return text
        except:
            return ""

    def parse_timesheet_entries(self, text, employee_name):
        """Parse timesheet entries from text"""
        entries = []
        if not text:
            return entries
            
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        for line in lines:
            # Look for date patterns
            date_patterns = [
                r'(\d{1,2}/\d{1,2}/\d{4})',
                r'(\d{1,2}-\d{1,2}-\d{4})',
            ]
            
            date_match = None
            for pattern in date_patterns:
                match = re.search(pattern, line)
                if match:
                    date_match = match
                    break
            
            if date_match:
                date = date_match.group(1)
                
                # Format date properly
                try:
                    if '/' in date:
                        parts = date.split('/')
                        if len(parts) == 3:
                            month, day, year = parts
                            formatted_date = f"{int(month):02d}/{int(day):02d}/{year}"
                    elif '-' in date:
                        parts = date.split('-')
                        if len(parts) == 3:
                            month, day, year = parts
                            formatted_date = f"{int(month):02d}/{int(day):02d}/{year}"
                    else:
                        formatted_date = date
                except:
                    formatted_date = date
                
                # Look for hours
                hours_patterns = [
                    r'(\d{1,2}(?:\.\d{1,2})?)\s*h(?:r|rs?|ours?)?',
                    r'(\d{1,2}(?:\.\d{1,2})?)\s+(?:Cost|Installation|Store|Enterprise|Product)',
                    r'\(\s*(\d{1,2}(?:\.\d{1,2})?)\s*\)',
                    r'(?:^|\s)(\d{1,2}(?:\.\d{1,2})?)(?=\s|$)'
                ]
                
                hours_match = None
                for pattern in hours_patterns:
                    match = re.search(pattern, line)
                    if match:
                        hours_match = match
                        break
                
                if hours_match:
                    try:
                        hours = float(hours_match.group(1))
                        if 0 <= hours <= 24:
                            entry = {
                                'Name': employee_name,
                                'Date': formatted_date,
                                'Hours': hours
                            }
                            entries.append(entry)
                    except:
                        pass
                else:
                    entry = {
                        'Name': employee_name,
                        'Date': formatted_date,
                        'Hours': "CHECK"
                    }
                    entries.append(entry)
        
        return entries

    def process_screenshot_from_bytes(self, image_bytes, consultant_name):
        """Process screenshot from bytes"""
        try:
            image = Image.open(io.BytesIO(image_bytes))
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Extract text
            text = self.extract_text_from_image(image)
            
            # Parse entries
            entries = self.parse_timesheet_entries(text, consultant_name)
            
            # Calculate total hours
            total_hours = sum(entry['Hours'] for entry in entries 
                            if isinstance(entry['Hours'], (int, float)))
            
            # Simulate system check
            system_hours = self.simulate_system_check(consultant_name)
            
            return {
                'consultant_name': consultant_name,
                'total_entries': len(entries),
                'screenshot_hours': total_hours,
                'system_hours': system_hours,
                'discrepancy_detected': total_hours != system_hours,
                'entries': entries,
                'status': 'success'
            }
            
        except Exception as e:
            return {'error': str(e), 'status': 'error'}

    def simulate_system_check(self, consultant_name):
        """Simulate system lookup"""
        demo_data = {
            'john smith': 38,
            'jane doe': 40,
            'mike johnson': 35,
            'sarah wilson': 40,
            'alex chen': 38
        }
        return demo_data.get(consultant_name.lower().strip(), 40)

processor = SimpleTimesheetProcessor()

@app.route('/')
def dashboard():
    return render_template_string('''
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>TimeVerify AI Dashboard</title>
        <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            
            body {
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                background: #f8f9fc;
                color: #1e1e2e;
                overflow-x: hidden;
            }
            
            .sidebar {
                position: fixed;
                left: 0;
                top: 0;
                width: 260px;
                height: 100vh;
                background: #ffffff;
                border-right: 1px solid #e3e6f0;
                z-index: 1000;
                transition: transform 0.3s ease;
            }
            
            .sidebar-header {
                padding: 24px 20px;
                border-bottom: 1px solid #e3e6f0;
                display: flex;
                align-items: center;
                gap: 12px;
            }
            
            .logo {
                width: 36px;
                height: 36px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                border-radius: 8px;
                display: flex;
                align-items: center;
                justify-content: center;
                color: white;
                font-size: 18px;
                font-weight: bold;
            }
            
            .logo-text {
                font-size: 20px;
                font-weight: 600;
                color: #1e1e2e;
            }
            
            .nav-menu {
                padding: 20px 0;
            }
            
            .nav-item {
                display: flex;
                align-items: center;
                padding: 12px 20px;
                color: #8b949e;
                text-decoration: none;
                transition: all 0.2s ease;
                border-left: 3px solid transparent;
            }
            
            .nav-item:hover, .nav-item.active {
                color: #667eea;
                background: #f8f9ff;
                border-left-color: #667eea;
            }
            
            .nav-item i {
                width: 20px;
                margin-right: 12px;
                font-size: 16px;
            }
            
            .main-content {
                margin-left: 260px;
                min-height: 100vh;
            }
            
            .header {
                background: white;
                padding: 20px 32px;
                border-bottom: 1px solid #e3e6f0;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }
            
            .header h1 {
                font-size: 28px;
                font-weight: 600;
                color: #1e1e2e;
            }
            
            .user-info {
                display: flex;
                align-items: center;
                gap: 12px;
                color: #8b949e;
            }
            
            .dashboard-grid {
                padding: 32px;
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
                gap: 24px;
            }
            
            .stats-row {
                grid-column: 1 / -1;
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
                gap: 24px;
                margin-bottom: 8px;
            }
            
            .stat-card {
                background: white;
                border-radius: 12px;
                padding: 24px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                position: relative;
                overflow: hidden;
            }
            
            .stat-card::before {
                content: '';
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                height: 4px;
            }
            
            .stat-card.processed::before { background: linear-gradient(90deg, #4facfe, #00f2fe); }
            .stat-card.hours::before { background: linear-gradient(90deg, #43e97b, #38f9d7); }
            .stat-card.accuracy::before { background: linear-gradient(90deg, #fa709a, #fee140); }
            .stat-card.savings::before { background: linear-gradient(90deg, #a8edea, #fed6e3); }
            
            .stat-header {
                display: flex;
                justify-content: space-between;
                align-items: flex-start;
                margin-bottom: 16px;
            }
            
            .stat-title {
                font-size: 14px;
                color: #8b949e;
                font-weight: 500;
            }
            
            .stat-icon {
                width: 40px;
                height: 40px;
                border-radius: 8px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 18px;
                color: white;
            }
            
            .stat-icon.processed { background: linear-gradient(135deg, #4facfe, #00f2fe); }
            .stat-icon.hours { background: linear-gradient(135deg, #43e97b, #38f9d7); }
            .stat-icon.accuracy { background: linear-gradient(135deg, #fa709a, #fee140); }
            .stat-icon.savings { background: linear-gradient(135deg, #a8edea, #fed6e3); }
            
            .stat-value {
                font-size: 32px;
                font-weight: 700;
                color: #1e1e2e;
                margin-bottom: 8px;
            }
            
            .stat-change {
                font-size: 12px;
                color: #22c55e;
                display: flex;
                align-items: center;
                gap: 4px;
            }
            
            .upload-card {
                background: white;
                border-radius: 12px;
                padding: 32px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                text-align: center;
                border: 2px dashed #e3e6f0;
                transition: all 0.3s ease;
            }
            
            .upload-card:hover {
                border-color: #667eea;
                background: #f8f9ff;
                transform: translateY(-2px);
                box-shadow: 0 4px 12px rgba(102, 126, 234, 0.15);
            }
            
            .upload-area {
                border: 2px dashed #d1d5db;
                border-radius: 8px;
                padding: 24px;
                margin-bottom: 20px;
                cursor: pointer;
                transition: all 0.3s ease;
                background: #f9fafb;
            }
            
            .upload-area:hover {
                border-color: #667eea;
                background: #f0f4ff;
            }
            
            .upload-area.dragover {
                border-color: #667eea;
                background: #e0e7ff;
                transform: scale(1.02);
            }
            
            .upload-icon {
                width: 64px;
                height: 64px;
                background: linear-gradient(135deg, #667eea, #764ba2);
                border-radius: 16px;
                display: flex;
                align-items: center;
                justify-content: center;
                color: white;
                font-size: 28px;
                margin: 0 auto 20px;
            }
            
            .upload-title {
                font-size: 20px;
                font-weight: 600;
                color: #1e1e2e;
                margin-bottom: 8px;
            }
            
            .upload-subtitle {
                color: #8b949e;
                margin-bottom: 24px;
            }
            
            .file-input {
                display: none;
            }
            
            .input-group {
                margin: 16px 0;
                text-align: left;
            }
            
            .input-group input {
                width: 100%;
                padding: 12px 16px;
                border: 2px solid #e3e6f0;
                border-radius: 8px;
                font-size: 14px;
                transition: border-color 0.2s ease;
            }
            
            .input-group input:focus {
                outline: none;
                border-color: #667eea;
            }
            
            .btn {
                background: linear-gradient(135deg, #667eea, #764ba2);
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: 500;
                cursor: pointer;
                transition: all 0.2s ease;
                display: inline-flex;
                align-items: center;
                gap: 8px;
            }
            
            .btn:hover {
                transform: translateY(-1px);
                box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
            }
            
            .btn-secondary {
                background: linear-gradient(135deg, #f093fb, #f5576c);
            }
            
            .processing-card {
                background: white;
                border-radius: 12px;
                padding: 24px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                display: none;
                grid-column: 1 / -1;
            }
            
            .processing-header {
                display: flex;
                align-items: center;
                gap: 12px;
                margin-bottom: 20px;
            }
            
            .spinner {
                width: 20px;
                height: 20px;
                border: 2px solid #e3e6f0;
                border-top: 2px solid #667eea;
                border-radius: 50%;
                animation: spin 1s linear infinite;
            }
            
            @keyframes spin {
                0% { transform: rotate(0deg); }
                100% { transform: rotate(360deg); }
            }
            
            .processing-steps {
                background: #f8f9fc;
                border-radius: 8px;
                padding: 16px;
                margin-top: 16px;
            }
            
            .processing-step {
                display: flex;
                align-items: center;
                padding: 8px 0;
                color: #8b949e;
                font-size: 14px;
                border-bottom: 1px solid #e3e6f0;
            }
            
            .processing-step:last-child {
                border-bottom: none;
            }
            
            .processing-step.active {
                color: #667eea;
                font-weight: 500;
            }
            
            .processing-step.completed {
                color: #22c55e;
            }
            
            .processing-step i {
                width: 20px;
                margin-right: 8px;
            }
            
            .results-card {
                background: white;
                border-radius: 12px;
                padding: 24px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                display: none;
                grid-column: 1 / -1;
            }
            
            .results-header {
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 24px;
            }
            
            .results-title {
                font-size: 18px;
                font-weight: 600;
                color: #1e1e2e;
            }
            
            .download-buttons {
                display: flex;
                gap: 8px;
            }
            
            .btn-small {
                padding: 8px 16px;
                font-size: 12px;
            }
            
            .status-badge {
                display: inline-flex;
                align-items: center;
                gap: 6px;
                padding: 6px 12px;
                border-radius: 6px;
                font-size: 12px;
                font-weight: 500;
            }
            
            .status-success {
                background: #dcfce7;
                color: #16a34a;
            }
            
            .status-warning {
                background: #fef3c7;
                color: #d97706;
            }
            
            .results-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 16px;
                margin: 20px 0;
            }
            
            .result-metric {
                text-align: center;
                padding: 16px;
                background: #f8f9fc;
                border-radius: 8px;
            }
            
            .result-value {
                font-size: 24px;
                font-weight: 700;
                color: #1e1e2e;
            }
            
            .result-label {
                font-size: 12px;
                color: #8b949e;
                margin-top: 4px;
            }
            
            /* Separate table section */
            .table-section {
                background: white;
                border-radius: 12px;
                padding: 24px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                margin: 32px;
                display: none;
            }
            
            .table-header {
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 20px;
                padding-bottom: 16px;
                border-bottom: 2px solid #e3e6f0;
            }
            
            .table-title {
                font-size: 20px;
                font-weight: 600;
                color: #1e1e2e;
                display: flex;
                align-items: center;
                gap: 8px;
            }
            
            .data-table {
                background: #f8f9fc;
                border-radius: 8px;
                padding: 16px;
                max-height: 400px;
                overflow-y: auto;
            }
            
            .data-table table {
                width: 100%;
                border-collapse: collapse;
                font-size: 13px;
            }
            
            .data-table th {
                background: #667eea;
                color: white;
                padding: 12px 8px;
                text-align: left;
                font-weight: 600;
                border: 1px solid #5a6fd8;
                position: sticky;
                top: 0;
                z-index: 10;
            }
            
            .data-table td {
                padding: 10px 8px;
                border: 1px solid #e3e6f0;
                background: white;
            }
            
            .data-table tr:hover td {
                background: #f8f9ff;
            }
            
            .check-value {
                color: #dc3545;
                font-weight: bold;
                background: #ffe6e6 !important;
            }
            
            .table-stats {
                margin-top: 16px;
                padding: 12px;
                background: #f0f4ff;
                border-radius: 6px;
                font-size: 12px;
                color: #667eea;
                display: flex;
                justify-content: space-between;
                flex-wrap: wrap;
                gap: 16px;
            }
            
            .table-stat {
                display: flex;
                align-items: center;
                gap: 4px;
            }
            
            @media (max-width: 768px) {
                .sidebar {
                    transform: translateX(-100%);
                }
                
                .main-content {
                    margin-left: 0;
                }
                
                .dashboard-grid {
                    padding: 16px;
                    grid-template-columns: 1fr;
                }
                
                .stats-row {
                    grid-template-columns: 1fr;
                }
                
                .table-section {
                    margin: 16px;
                }
            }
        </style>
    </head>
    <body>
        <div class="sidebar">
            <div class="sidebar-header">
                <div class="logo">T</div>
                <div class="logo-text">TimeVerify</div>
            </div>
            <nav class="nav-menu">
                <a href="#" class="nav-item active">
                    <i class="fas fa-chart-line"></i>
                    Dashboard
                </a>
                <a href="#" class="nav-item">
                    <i class="fas fa-upload"></i>
                    Upload
                </a>
                <a href="#" class="nav-item">
                    <i class="fas fa-file-alt"></i>
                    Reports
                </a>
                <a href="#" class="nav-item">
                    <i class="fas fa-cog"></i>
                    Settings
                </a>
            </nav>
        </div>
        
        <div class="main-content">
            <div class="header">
                <h1>Dashboard</h1>
                <div class="user-info">
                    <span>Last updated: <span id="lastUpdate">Never</span></span>
                    <div style="width: 2px; height: 20px; background: #e3e6f0;"></div>
                    <i class="fas fa-user-circle" style="font-size: 24px; color: #667eea;"></i>
                </div>
            </div>
            
            <div class="dashboard-grid">
                <div class="stats-row">
                    <div class="stat-card processed">
                        <div class="stat-header">
                            <div class="stat-title">Documents Processed</div>
                            <div class="stat-icon processed">
                                <i class="fas fa-file-check"></i>
                            </div>
                        </div>
                        <div class="stat-value" id="processedCount">0</div>
                        <div class="stat-change">
                            <i class="fas fa-arrow-up"></i>
                            <span>Today</span>
                        </div>
                    </div>
                    
                    <div class="stat-card hours">
                        <div class="stat-header">
                            <div class="stat-title">Hours Extracted</div>
                            <div class="stat-icon hours">
                                <i class="fas fa-clock"></i>
                            </div>
                        </div>
                        <div class="stat-value" id="hoursCount">0</div>
                        <div class="stat-change">
                            <i class="fas fa-arrow-up"></i>
                            <span>This session</span>
                        </div>
                    </div>
                    
                    <div class="stat-card accuracy">
                        <div class="stat-header">
                            <div class="stat-title">Accuracy Rate</div>
                            <div class="stat-icon accuracy">
                                <i class="fas fa-bullseye"></i>
                            </div>
                        </div>
                        <div class="stat-value">95%</div>
                        <div class="stat-change">
                            <i class="fas fa-arrow-up"></i>
                            <span>OCR Quality</span>
                        </div>
                    </div>
                    
                    <div class="stat-card savings">
                        <div class="stat-header">
                            <div class="stat-title">Time Savings</div>
                            <div class="stat-icon savings">
                                <i class="fas fa-stopwatch"></i>
                            </div>
                        </div>
                        <div class="stat-value" id="timeSaved">0min</div>
                        <div class="stat-change">
                            <i class="fas fa-arrow-up"></i>
                            <span>vs Manual</span>
                        </div>
                    </div>
                </div>
                
                <div class="upload-card">
                    <div class="upload-icon">
                        <i class="fas fa-camera"></i>
                    </div>
                    <div class="upload-title">Upload Screenshot</div>
                    <div class="upload-subtitle">PNG, JPG files supported</div>
                    <form id="screenshotForm" enctype="multipart/form-data">
                        <div class="upload-area" onclick="document.getElementById('screenshotFile').click()">
                            <i class="fas fa-cloud-upload-alt" style="font-size: 24px; color: #667eea; margin-bottom: 8px;"></i>
                            <p style="margin: 0; color: #6b7280;">Click to select file or drag & drop</p>
                            <p id="screenshotFileName" style="margin: 8px 0 0 0; font-size: 12px; color: #667eea; font-weight: 500;"></p>
                        </div>
                        <input type="file" id="screenshotFile" name="screenshot" accept="image/*" class="file-input" required>
                        <div class="input-group">
                            <input type="text" name="consultant_name" placeholder="Consultant Name" required onclick="event.stopPropagation();">
                        </div>
                        <button type="submit" class="btn">
                            <i class="fas fa-upload"></i>
                            Process Screenshot
                        </button>
                    </form>
                </div>
                
                <div class="upload-card">
                    <div class="upload-icon">
                        <i class="fas fa-file-word"></i>
                    </div>
                    <div class="upload-title">Upload Document</div>
                    <div class="upload-subtitle">DOCX, DOC files supported</div>
                    <form id="documentForm" enctype="multipart/form-data">
                        <div class="upload-area" onclick="document.getElementById('documentFile').click()">
                            <i class="fas fa-file-upload" style="font-size: 24px; color: #667eea; margin-bottom: 8px;"></i>
                            <p style="margin: 0; color: #6b7280;">Click to select file or drag & drop</p>
                            <p id="documentFileName" style="margin: 8px 0 0 0; font-size: 12px; color: #667eea; font-weight: 500;"></p>
                        </div>
                        <input type="file" id="documentFile" name="document" accept=".docx,.doc" class="file-input" required>
                        <div class="input-group">
                            <input type="text" name="consultant_name" placeholder="Consultant Name (optional)" onclick="event.stopPropagation();">
                        </div>
                        <button type="submit" class="btn btn-secondary">
                            <i class="fas fa-file-upload"></i>
                            Extract & Process
                        </button>
                    </form>
                </div>
                
                <div class="upload-card">
                    <div class="upload-icon">
                        <i class="fas fa-folder-open"></i>
                    </div>
                    <div class="upload-title">Bulk Upload</div>
                    <div class="upload-subtitle">Multiple DOCX files at once</div>
                    <form id="bulkForm" enctype="multipart/form-data">
                        <div class="upload-area" onclick="document.getElementById('bulkFiles').click()">
                            <i class="fas fa-files" style="font-size: 24px; color: #667eea; margin-bottom: 8px;"></i>
                            <p style="margin: 0; color: #6b7280;">Select multiple documents</p>
                            <p id="bulkFileName" style="margin: 8px 0 0 0; font-size: 12px; color: #667eea; font-weight: 500;"></p>
                        </div>
                        <input type="file" id="bulkFiles" name="documents" accept=".docx,.doc" class="file-input" multiple required>
                        <button type="submit" class="btn" style="background: linear-gradient(135deg, #43e97b, #38f9d7);">
                            <i class="fas fa-cogs"></i>
                            Process All Documents
                        </button>
                    </form>
                </div>
                
                <div class="processing-card" id="processingCard">
                    <div class="processing-header">
                        <div class="spinner"></div>
                        <span id="processingTitle">Processing your request...</span>
                    </div>
                    <div class="processing-steps">
                        <div class="processing-step" id="step1">
                            <i class="fas fa-circle-notch"></i>
                            <span id="step1Text">Initializing...</span>
                        </div>
                        <div class="processing-step" id="step2">
                            <i class="fas fa-circle-notch"></i>
                            <span id="step2Text">Waiting...</span>
                        </div>
                        <div class="processing-step" id="step3">
                            <i class="fas fa-circle-notch"></i>
                            <span id="step3Text">Waiting...</span>
                        </div>
                        <div class="processing-step" id="step4">
                            <i class="fas fa-circle-notch"></i>
                            <span id="step4Text">Waiting...</span>
                        </div>
                        <div class="processing-step" id="step5">
                            <i class="fas fa-circle-notch"></i>
                            <span id="step5Text">Waiting...</span>
                        </div>
                    </div>
                </div>
                
                <div class="results-card" id="resultsCard">
                    <div class="results-header">
                        <div class="results-title">Processing Results</div>
                        <div class="download-buttons">
                            <button onclick="downloadData('excel')" class="btn btn-small">
                                <i class="fas fa-file-excel"></i>
                                Excel
                            </button>
                            <button onclick="downloadData('json')" class="btn btn-small btn-secondary">
                                <i class="fas fa-download"></i>
                                JSON
                            </button>
                        </div>
                    </div>
                    <div id="resultsStatus"></div>
                    <div class="results-grid" id="resultsMetrics"></div>
                </div>
            </div>
            
            <!-- Separate Table Section -->
            <div class="table-section" id="tableSection">
                <div class="table-header">
                    <div class="table-title">
                        <i class="fas fa-table"></i>
                        Extracted Timesheet Data
                    </div>
                    <div class="download-buttons">
                        <button onclick="downloadData('excel')" class="btn btn-small">
                            <i class="fas fa-file-excel"></i>
                            Export Excel
                        </button>
                        <button onclick="downloadData('json')" class="btn btn-small btn-secondary">
                            <i class="fas fa-download"></i>
                            Export JSON
                        </button>
                    </div>
                </div>
                <div class="data-table" id="dataTable"></div>
            </div>
        </div>
        
        <script>
            let currentResults = null;
            let sessionStats = {
                processed: 0,
                hours: 0,
                timeSaved: 0
            };
            
            // Form handlers
            document.getElementById('screenshotForm').onsubmit = async function(e) {
                e.preventDefault();
                const fileName = document.getElementById('screenshotFile').files[0]?.name || 'screenshot';
                await processForm(this, '/api/process-screenshot', fileName);
            };
            
            document.getElementById('documentForm').onsubmit = async function(e) {
                e.preventDefault();
                const fileName = document.getElementById('documentFile').files[0]?.name || 'document';
                await processForm(this, '/api/process-document', fileName);
            };
            
            document.getElementById('bulkForm').onsubmit = async function(e) {
                e.preventDefault();
                const files = document.getElementById('bulkFiles').files;
                const fileNames = Array.from(files).map(f => f.name);
                await processForm(this, '/api/process-bulk', fileNames);
            };
            
            // File input handlers with filename display
            document.getElementById('screenshotFile').onchange = function() {
                const fileName = this.files[0] ? this.files[0].name : '';
                document.getElementById('screenshotFileName').textContent = fileName ? `Selected: ${fileName}` : '';
            };
            
            document.getElementById('documentFile').onchange = function() {
                const fileName = this.files[0] ? this.files[0].name : '';
                document.getElementById('documentFileName').textContent = fileName ? `Selected: ${fileName}` : '';
            };
            
            document.getElementById('bulkFiles').onchange = function() {
                const fileCount = this.files.length;
                if (fileCount > 0) {
                    document.getElementById('bulkFileName').textContent = `Selected: ${fileCount} file(s)`;
                } else {
                    document.getElementById('bulkFileName').textContent = '';
                }
            };
            
            // Enhanced drag and drop functionality
            setupDragAndDrop('screenshotFile', document.querySelector('#screenshotForm .upload-area'));
            setupDragAndDrop('documentFile', document.querySelector('#documentForm .upload-area'));
            setupDragAndDrop('bulkFiles', document.querySelector('#bulkForm .upload-area'));
            
            function setupDragAndDrop(inputId, dropArea) {
                const input = document.getElementById(inputId);
                
                ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
                    dropArea.addEventListener(eventName, preventDefaults, false);
                });
                
                function preventDefaults(e) {
                    e.preventDefault();
                    e.stopPropagation();
                }
                
                ['dragenter', 'dragover'].forEach(eventName => {
                    dropArea.addEventListener(eventName, () => dropArea.classList.add('dragover'), false);
                });
                
                ['dragleave', 'drop'].forEach(eventName => {
                    dropArea.addEventListener(eventName, () => dropArea.classList.remove('dragover'), false);
                });
                
                dropArea.addEventListener('drop', function(e) {
                    const files = e.dataTransfer.files;
                    if (files.length > 0) {
                        input.files = files;
                        input.dispatchEvent(new Event('change'));
                    }
                });
            }
            
            async function processForm(form, endpoint, fileNames) {
                showProcessing(fileNames);
                
                try {
                    const formData = new FormData(form);
                    const response = await fetch(endpoint, {
                        method: 'POST',
                        body: formData
                    });
                    
                    const result = await response.json();
                    
                    if (result.error) {
                        throw new Error(result.error);
                    }
                    
                    currentResults = result;
                    updateSessionStats(result);
                    showResults(result);
                    showTable(result);
                    updateLastUpdated();
                    
                } catch (error) {
                    showError(error.message);
                } finally {
                    hideProcessing();
                }
            }
            
            function showProcessing(fileNames) {
                document.getElementById('processingCard').style.display = 'block';
                document.getElementById('resultsCard').style.display = 'none';
                document.getElementById('tableSection').style.display = 'none';
                
                // Reset all steps
                for (let i = 1; i <= 5; i++) {
                    const step = document.getElementById(`step${i}`);
                    step.className = 'processing-step';
                    step.querySelector('i').className = 'fas fa-circle-notch';
                }
                
                // Update title based on file type
                let title = 'Processing your request...';
                if (Array.isArray(fileNames)) {
                    title = `Processing ${fileNames.length} files...`;
                    document.getElementById('processingTitle').textContent = title;
                    simulateBulkProcessing(fileNames);
                } else {
                    title = `Processing ${fileNames}...`;
                    document.getElementById('processingTitle').textContent = title;
                    simulateProcessing(fileNames);
                }
            }
            
            function simulateProcessing(fileName) {
                const steps = [
                    { id: 'step1', text: `Reading file: ${fileName}`, delay: 500 },
                    { id: 'step2', text: 'Extracting images and text...', delay: 1000 },
                    { id: 'step3', text: 'Running OCR analysis...', delay: 1500 },
                    { id: 'step4', text: 'Parsing timesheet entries...', delay: 2000 },
                    { id: 'step5', text: 'Validating and generating results...', delay: 2500 }
                ];
                
                steps.forEach((step, index) => {
                    setTimeout(() => {
                        const stepElement = document.getElementById(step.id);
                        stepElement.className = 'processing-step active';
                        stepElement.querySelector('i').className = 'fas fa-spinner fa-spin';
                        document.getElementById(`${step.id}Text`).textContent = step.text;
                        
                        // Mark previous steps as completed
                        for (let i = 1; i < index + 1; i++) {
                            const prevStep = document.getElementById(`step${i}`);
                            if (prevStep && i < index + 1) {
                                prevStep.className = 'processing-step completed';
                                prevStep.querySelector('i').className = 'fas fa-check-circle';
                            }
                        }
                    }, step.delay);
                });
            }
            
            function simulateBulkProcessing(fileNames) {
                const steps = [
                    { id: 'step1', text: `Preparing ${fileNames.length} documents...`, delay: 300 },
                    { id: 'step2', text: `Processing: ${fileNames[0] || 'first file'}`, delay: 800 },
                    { id: 'step3', text: 'Running OCR on multiple images...', delay: 1500 },
                    { id: 'step4', text: `Processing: ${fileNames[1] || 'additional files'}...`, delay: 2200 },
                    { id: 'step5', text: 'Consolidating all results...', delay: 3000 }
                ];
                
                // Show more files being processed if available
                if (fileNames.length > 2) {
                    steps[3].text = `Processing remaining ${fileNames.length - 1} files...`;
                }
                
                steps.forEach((step, index) => {
                    setTimeout(() => {
                        const stepElement = document.getElementById(step.id);
                        stepElement.className = 'processing-step active';
                        stepElement.querySelector('i').className = 'fas fa-spinner fa-spin';
                        document.getElementById(`${step.id}Text`).textContent = step.text;
                        
                        // Mark previous steps as completed
                        for (let i = 1; i < index + 1; i++) {
                            const prevStep = document.getElementById(`step${i}`);
                            if (prevStep && i < index + 1) {
                                prevStep.className = 'processing-step completed';
                                prevStep.querySelector('i').className = 'fas fa-check-circle';
                            }
                        }
                    }, step.delay);
                });
            }
            
            function hideProcessing() {
                document.getElementById('processingCard').style.display = 'none';
            }
            
            function showResults(result) {
                const resultsCard = document.getElementById('resultsCard');
                const statusDiv = document.getElementById('resultsStatus');
                const metricsDiv = document.getElementById('resultsMetrics');
                
                // Check if this is bulk processing results
                if (result.results && Array.isArray(result.results)) {
                    // Bulk processing results
                    showBulkResults(result);
                    return;
                }
                
                // Single file results
                const hasDiscrepancy = result.discrepancy_detected;
                statusDiv.innerHTML = `
                    <div class="status-badge ${hasDiscrepancy ? 'status-warning' : 'status-success'}">
                        <i class="fas ${hasDiscrepancy ? 'fa-exclamation-triangle' : 'fa-check-circle'}"></i>
                        ${hasDiscrepancy ? 'Discrepancy Detected' : 'All Verified'}
                    </div>
                `;
                
                // Metrics
                metricsDiv.innerHTML = `
                    <div class="result-metric">
                        <div class="result-value">${result.consultant_name || 'N/A'}</div>
                        <div class="result-label">Consultant</div>
                    </div>
                    <div class="result-metric">
                        <div class="result-value">${result.total_entries || 0}</div>
                        <div class="result-label">Entries Found</div>
                    </div>
                    <div class="result-metric">
                        <div class="result-value">${result.screenshot_hours || 0}</div>
                        <div class="result-label">Screenshot Hours</div>
                    </div>
                    <div class="result-metric">
                        <div class="result-value">${result.system_hours || 0}</div>
                        <div class="result-label">System Hours</div>
                    </div>
                `;
                
                resultsCard.style.display = 'block';
            }
            
            function showBulkResults(result) {
                const resultsCard = document.getElementById('resultsCard');
                const statusDiv = document.getElementById('resultsStatus');
                const metricsDiv = document.getElementById('resultsMetrics');
                
                const summary = result.summary || {};
                const discrepancies = summary.discrepancies_found || 0;
                
                // Status for bulk processing
                statusDiv.innerHTML = `
                    <div class="status-badge ${discrepancies > 0 ? 'status-warning' : 'status-success'}">
                        <i class="fas ${discrepancies > 0 ? 'fa-exclamation-triangle' : 'fa-check-circle'}"></i>
                        Bulk Processing: ${discrepancies > 0 ? `${discrepancies} Discrepancies Found` : 'All Verified'}
                    </div>
                `;
                
                // Bulk metrics
                metricsDiv.innerHTML = `
                    <div class="result-metric">
                        <div class="result-value">${summary.successful_documents || 0}</div>
                        <div class="result-label">Documents Processed</div>
                    </div>
                    <div class="result-metric">
                        <div class="result-value">${summary.total_images || 0}</div>
                        <div class="result-label">Images Processed</div>
                    </div>
                    <div class="result-metric">
                        <div class="result-value">${summary.total_entries || 0}</div>
                        <div class="result-label">Total Entries</div>
                    </div>
                    <div class="result-metric">
                        <div class="result-value">${discrepancies}</div>
                        <div class="result-label">Discrepancies</div>
                    </div>
                `;
                
                resultsCard.style.display = 'block';
            }
            
            function showTable(result) {
                const tableSection = document.getElementById('tableSection');
                const dataTable = document.getElementById('dataTable');
                
                let entries = [];
                let isMultipleFiles = false;
                
                // Get entries based on result type
                if (result.results && Array.isArray(result.results)) {
                    // Bulk processing - combine all entries
                    entries = result.entries || [];
                    isMultipleFiles = true;
                } else {
                    // Single file processing
                    entries = result.entries || [];
                }
                
                if (entries.length === 0) {
                    dataTable.innerHTML = '<p style="text-align: center; color: #8b949e; padding: 40px;">No timesheet entries found</p>';
                    tableSection.style.display = 'block';
                    return;
                }
                
                // Build table HTML
                let tableHTML = `
                    <table>
                        <thead>
                            <tr>
                                <th>Name</th>
                                <th>Date</th>
                                <th>Hours</th>
                                ${isMultipleFiles ? '<th>Source</th>' : ''}
                            </tr>
                        </thead>
                        <tbody>
                `;
                
                entries.forEach((entry, index) => {
                    const hoursClass = entry.Hours === "CHECK" ? "check-value" : "";
                    const sourceFile = isMultipleFiles ? (entry.source_file || 'Unknown') : '';
                    
                    tableHTML += `
                        <tr>
                            <td>${entry.Name || 'N/A'}</td>
                            <td>${entry.Date || 'N/A'}</td>
                            <td class="${hoursClass}">${entry.Hours !== undefined ? entry.Hours : 'N/A'}</td>
                            ${isMultipleFiles ? `<td style="font-size: 11px;">${sourceFile}</td>` : ''}
                        </tr>
                    `;
                });
                
                tableHTML += `
                        </tbody>
                    </table>
                `;
                
                // Add statistics
                const validHours = entries.filter(e => typeof e.Hours === 'number').length;
                const checkEntries = entries.filter(e => e.Hours === 'CHECK').length;
                const totalHours = entries.reduce((sum, e) => sum + (typeof e.Hours === 'number' ? e.Hours : 0), 0);
                
                tableHTML += `
                    <div class="table-stats">
                        <div class="table-stat">
                            <i class="fas fa-list"></i>
                            <span>Total Entries: ${entries.length}</span>
                        </div>
                        <div class="table-stat">
                            <i class="fas fa-check-circle"></i>
                            <span>Valid Hours: ${validHours}</span>
                        </div>
                        <div class="table-stat">
                            <i class="fas fa-exclamation-triangle"></i>
                            <span>Needs Review: ${checkEntries}</span>
                        </div>
                        <div class="table-stat">
                            <i class="fas fa-clock"></i>
                            <span>Total Hours: ${totalHours}</span>
                        </div>
                    </div>
                `;
                
                dataTable.innerHTML = tableHTML;
                tableSection.style.display = 'block';
            }
            
            function showError(message) {
                document.getElementById('resultsCard').style.display = 'block';
                document.getElementById('resultsStatus').innerHTML = `
                    <div class="status-badge status-warning">
                        <i class="fas fa-exclamation-triangle"></i>
                        Error: ${message}
                    </div>
                `;
                document.getElementById('resultsMetrics').innerHTML = '';
            }
            
            function updateSessionStats(result) {
                sessionStats.processed++;
                
                if (result.results && Array.isArray(result.results)) {
                    // Bulk processing
                    sessionStats.hours += result.summary?.total_entries || 0;
                    sessionStats.timeSaved += (result.results.length * 10);
                } else {
                    // Single file
                    sessionStats.hours += result.screenshot_hours || 0;
                    sessionStats.timeSaved += 10;
                }
                
                document.getElementById('processedCount').textContent = sessionStats.processed;
                document.getElementById('hoursCount').textContent = sessionStats.hours;
                document.getElementById('timeSaved').textContent = sessionStats.timeSaved + 'min';
            }
            
            function updateLastUpdated() {
                const now = new Date();
                document.getElementById('lastUpdate').textContent = now.toLocaleTimeString();
            }
            
            async function downloadData(format) {
                if (!currentResults) return;
                
                try {
                    if (format === 'excel') {
                        // Send data to backend for proper Excel generation
                        const response = await fetch('/api/download/excel', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify(currentResults)
                        });
                        
                        if (response.ok) {
                            const blob = await response.blob();
                            const url = URL.createObjectURL(blob);
                            const a = document.createElement('a');
                            a.href = url;
                            a.download = `timesheet_${new Date().toISOString().split('T')[0]}.xlsx`;
                            document.body.appendChild(a);
                            a.click();
                            document.body.removeChild(a);
                            URL.revokeObjectURL(url);
                        } else {
                            throw new Error('Excel generation failed');
                        }
                        
                    } else if (format === 'json') {
                        const jsonData = JSON.stringify(currentResults, null, 2);
                        const blob = new Blob([jsonData], { type: 'application/json' });
                        const filename = `timesheet_${new Date().toISOString().split('T')[0]}.json`;
                        
                        const url = URL.createObjectURL(blob);
                        const a = document.createElement('a');
                        a.href = url;
                        a.download = filename;
                        document.body.appendChild(a);
                        a.click();
                        document.body.removeChild(a);
                        URL.revokeObjectURL(url);
                    }
                    
                } catch (error) {
                    alert('Download failed: ' + error.message);
                }
            }
            
            // Initialize dashboard
            updateLastUpdated();
        </script>
    </body>
    </html>
    ''')

@app.route('/api/process-screenshot', methods=['POST'])
def process_screenshot():
    try:
        if 'screenshot' not in request.files:
            return jsonify({'error': 'No screenshot file provided'}), 400
        
        file = request.files['screenshot']
        consultant_name = request.form.get('consultant_name', 'Unknown')
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        print(f"Processing screenshot for: {consultant_name}")
        
        # Process using simplified logic
        result = processor.process_screenshot_from_bytes(file.read(), consultant_name)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/process-document', methods=['POST'])
def process_document():
    try:
        if 'document' not in request.files:
            return jsonify({'error': 'No document file provided'}), 400
        
        file = request.files['document']
        consultant_name = request.form.get('consultant_name', '')
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
            file.save(temp_file.name)
            temp_path = temp_file.name
        
        try:
            # Extract consultant name from filename if not provided
            if not consultant_name:
                consultant_name = processor.extract_name_from_filename(file.filename)
            
            print(f"Processing document: {file.filename} for consultant: {consultant_name}")
            
            # Extract images from Word document
            images = processor.extract_images_from_word_file(temp_path)
            
            if not images:
                return jsonify({
                    'error': 'No images found in the document',
                    'consultant_name': consultant_name,
                    'filename': file.filename
                })
            
            # Process each image with OCR
            all_entries = []
            
            for idx, image in enumerate(images, 1):
                print(f"Processing image {idx}/{len(images)}")
                
                # Extract text using OCR
                text = processor.extract_text_from_image(image)
                
                # Parse entries
                entries = processor.parse_timesheet_entries(text, consultant_name)
                all_entries.extend(entries)
            
            # Calculate totals
            total_hours = sum(entry['Hours'] for entry in all_entries 
                            if isinstance(entry['Hours'], (int, float)))
            
            # Simulate system check
            system_hours = processor.simulate_system_check(consultant_name)
            
            result = {
                'consultant_name': consultant_name,
                'filename': file.filename,
                'total_images': len(images),
                'total_entries': len(all_entries),
                'screenshot_hours': total_hours,
                'system_hours': system_hours,
                'discrepancy_detected': total_hours != system_hours,
                'entries': all_entries,
                'status': 'success'
            }
            
            print(f"Processing complete: {total_hours} hours extracted from {len(images)} images")
            return jsonify(result)
            
        finally:
            # Clean up temporary file
            os.unlink(temp_path)
            
    except Exception as e:
        print(f"Error processing document: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/process-bulk', methods=['POST'])
def process_bulk():
    try:
        if 'documents' not in request.files:
            return jsonify({'error': 'No document files provided'}), 400
        
        files = request.files.getlist('documents')
        
        if not files:
            return jsonify({'error': 'No files selected'}), 400
        
        print(f"Processing {len(files)} documents in bulk...")
        
        all_results = []
        total_images = 0
        total_entries = 0
        all_entries = []  # For combined Excel export
        
        for file in files:
            if file.filename == '':
                continue
                
            try:
                # Save uploaded file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
                    file.save(temp_file.name)
                    temp_path = temp_file.name
                
                try:
                    # Extract consultant name from filename
                    consultant_name = processor.extract_name_from_filename(file.filename)
                    
                    print(f"Processing: {file.filename} for {consultant_name}")
                    
                    # Extract images from Word document
                    images = processor.extract_images_from_word_file(temp_path)
                    
                    if not images:
                        all_results.append({
                            'consultant_name': consultant_name,
                            'filename': file.filename,
                            'status': 'No images found',
                            'error': 'Document contains no embedded images',
                            'images_processed': 0,
                            'screenshot_hours': 0,
                            'system_hours': 0,
                            'discrepancy_detected': False
                        })
                        continue
                    
                    # Process all images in this document
                    file_entries = []
                    for image in images:
                        text = processor.extract_text_from_image(image)
                        entries = processor.parse_timesheet_entries(text, consultant_name)
                        # Add source file info to each entry
                        for entry in entries:
                            entry['source_file'] = file.filename
                        file_entries.extend(entries)
                    
                    # Calculate hours for this consultant
                    consultant_hours = sum(entry['Hours'] for entry in file_entries 
                                         if isinstance(entry['Hours'], (int, float)))
                    
                    system_hours = processor.simulate_system_check(consultant_name)
                    
                    file_result = {
                        'consultant_name': consultant_name,
                        'filename': file.filename,
                        'images_processed': len(images),
                        'entries_found': len(file_entries),
                        'screenshot_hours': consultant_hours,
                        'system_hours': system_hours,
                        'discrepancy_detected': consultant_hours != system_hours,
                        'status': 'Processed successfully',
                        'entries': file_entries  # Include entries for this file
                    }
                    
                    all_results.append(file_result)
                    all_entries.extend(file_entries)  # Add to combined entries
                    total_images += len(images)
                    total_entries += len(file_entries)
                    
                finally:
                    # Clean up temporary file
                    os.unlink(temp_path)
                    
            except Exception as e:
                all_results.append({
                    'consultant_name': processor.extract_name_from_filename(file.filename),
                    'filename': file.filename,
                    'status': 'Processing failed',
                    'error': str(e),
                    'images_processed': 0,
                    'screenshot_hours': 0,
                    'system_hours': 0,
                    'discrepancy_detected': False
                })
        
        # Summary statistics
        successful_files = [r for r in all_results if r.get('status') == 'Processed successfully']
        discrepancies = [r for r in successful_files if r.get('discrepancy_detected')]
        
        summary = {
            'total_documents': len(files),
            'successful_documents': len(successful_files),
            'total_images': total_images,
            'total_entries': total_entries,
            'discrepancies_found': len(discrepancies),
            'processing_time': f"{len(files) * 2.5:.1f} seconds",
            'manual_equivalent': f"{len(files) * 15} minutes"
        }
        
        # Prepare response with combined entries for Excel export
        response = {
            'summary': summary,
            'results': all_results,
            'entries': all_entries,  # Combined entries from all files
            'total_images': total_images,
            'total_entries': total_entries,
            'processing_timestamp': datetime.now().isoformat(),
            'bulk_processing': True
        }
        
        print(f"Bulk processing complete: {len(successful_files)}/{len(files)} files processed successfully")
        return jsonify(response)
        
    except Exception as e:
        print(f"Error in bulk processing: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/excel', methods=['POST'])
def download_excel():
    try:
        data = request.json
        
        # Prepare entries data
        entries = data.get('entries', [])
        if not entries:
            # Create single entry from main data
            entries = [{
                'Name': data.get('consultant_name', 'Unknown'),
                'Date': 'N/A',
                'Hours': data.get('screenshot_hours', 0)
            }]
        
        # Create Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Timesheet Data"
            
            # Headers
            headers = ['Name', 'Date', 'Hours']
            if any('source_file' in entry for entry in entries):
                headers.append('Source File')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Data rows
            for row_idx, entry in enumerate(entries, 2):
                name_value = entry.get('Name', '')
                date_value = entry.get('Date', '')
                hours_value = entry.get('Hours', '')
                source_value = entry.get('source_file', '')
                
                # Name column (A)
                cell = ws.cell(row=row_idx, column=1, value=name_value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Date column (B)
                cell = ws.cell(row=row_idx, column=2, value=date_value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Hours column (C)
                cell = ws.cell(row=row_idx, column=3, value=hours_value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Source file column (D) if present
                if 'Source File' in headers:
                    cell = ws.cell(row=row_idx, column=4, value=source_value)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                
                # Highlight CHECK entries
                if hours_value == "CHECK":
                    hours_cell = ws.cell(row=row_idx, column=3)
                    hours_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    hours_cell.font = Font(color="FF0000")
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 25  # Name
            ws.column_dimensions['B'].width = 15  # Date  
            ws.column_dimensions['C'].width = 10  # Hours
            if 'Source File' in headers:
                ws.column_dimensions['D'].width = 30  # Source File
            
            # Add filter
            filter_range = f"A1:{chr(ord('A') + len(headers) - 1)}{len(entries) + 1}"
            ws.auto_filter.ref = filter_range
            
            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'timesheet_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            
        except ImportError:
            # Fallback: Create proper CSV with Excel-friendly format
            output = io.StringIO()
            
            # Write BOM for Excel UTF-8 compatibility
            output.write('\ufeff')
            
            # Write headers
            headers = ['Name', 'Date', 'Hours']
            if any('source_file' in entry for entry in entries):
                headers.append('Source File')
            output.write(','.join(headers) + '\n')
            
            # Write data
            for entry in entries:
                name = str(entry.get('Name', '')).replace('"', '""')
                date = str(entry.get('Date', '')).replace('"', '""')
                hours = str(entry.get('Hours', ''))
                source = str(entry.get('source_file', '')).replace('"', '""')
                
                row = f'"{name}","{date}","{hours}"'
                if 'Source File' in headers:
                    row += f',"{source}"'
                output.write(row + '\n')
            
            # Create response
            response_data = output.getvalue()
            output.close()
            
            response = app.response_class(
                response=response_data,
                mimetype='text/csv',
                headers={"Content-disposition": f"attachment; filename=timesheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
            )
            return response
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health')
def health():
    return jsonify({
        'status': 'healthy',
        'service': 'TimeVerify AI Dashboard',
        'timestamp': datetime.now().isoformat()
    })

if __name__ == '__main__':
    import os
    # OpenShift compatibility - use PORT environment variable
    port = int(os.environ.get('PORT', 8080))
    host = os.environ.get('HOST', '0.0.0.0')
    
    print(f"🚀 TimeVerify AI Dashboard - Starting on {host}:{port}...")
    print(f"🌐 Server running on {host}:{port}")
    print("📊 Features: Modern Dashboard + Simplified Processing")
    
    app.run(host=host, port=port, debug=False)